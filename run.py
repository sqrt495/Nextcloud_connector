import os
import openpyxl
from datetime import datetime, timedelta
import shutil
import win32com.client
import config, selenium_with_nextcloud, unzip_and_sort, secret_keys, divide_versions, telegram_actions
from time import sleep
import stat


# tech defs:
def nowtime_str():
    return nowtime.strftime("%d.%m.%Y-%H.%M")

def log_with_nowtime_str(you_massage):
    print(nowtime_str(),
          you_massage,
          sep='\n')

# check dirs architecture
config.rise_up_project_architecture(temp_dir=config.temp_dir,
                                 archive_dir=config.archive_dir,
                                 cloud_dir=config.cloud_dir,
                                 sort_dir=config.sort_dir,
                                 tables_dir=config.tables_dir,
                                 control_mo_tables_dir=config.control_mo_tables_dir,
                                 delta_tables_dir=config.delta_tables_dir,
                                 new_patients_tables_dir=config.new_patients_tables_dir,
                                 not_table_kjz_dir=config.not_table_kjz_dir,
                                 others_files_dir=config.others_files_dir)

#  TODO: удаление отработанных файлов и очистку папок

# key dates:
nowtime = datetime.now()
today = datetime.now().strftime('%d.%m.%Y')
tomorrow = (datetime.now() + timedelta(days=1)).strftime('%d.%m.%Y')
day_after_tomorrow = (datetime.now() + timedelta(days=2)).strftime('%d.%m.%Y')
two_day_after_tomorrow = (datetime.now() + timedelta(days=3)).strftime('%d.%m.%Y')
today_weekday = datetime.today().weekday()

alph = 'АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ'

not_table_file_types = ['pdf', 'jpeg', 'jpg', 'png', 'zip', 'PDF', 'JPEG', 'JPG', 'PNG', 'ZIP']
arch_file_types = ['zip', 'ZIP']
table_file_types = ['xlsx', 'xlsm', 'xls', 'ods', 'XLSX', 'XLSM', 'XLS', 'ODS']

dictionary = unzip_and_sort.check_handler_dictionary

########################################################################################################################
# hello massege for admins
telegram_actions.send_message_for_users_by_list('Обработка файлов "Контроль МО" начата!')
log_with_nowtime_str('- = START = -')
# get files from cloud
selenium_with_nextcloud.download_clear_create_folder(
                             url=secret_keys.nextcloud_provider_url,
                             path_for_download=config.temp_dir,
                             folders_name_list=[today, tomorrow],
                             password='')

saved_archive = os.listdir(config.temp_dir)[0]
original_folder_name = os.listdir(config.temp_dir)[0].replace('.zip','')

shutil.copyfile(os.path.join(config.temp_dir,
                             saved_archive),
                os.path.join(config.cloud_dir,
                             nowtime_str() + '-' + saved_archive))

archive_to_cloud = f"{os.path.join(config.cloud_dir,nowtime_str() + '-' + saved_archive)}"

# upload archive
selenium_with_nextcloud.upload(url=secret_keys.nextcloud_upload_archive_url,
                               files=[archive_to_cloud],
                               password=secret_keys.nextcloud_upload_archive_url_pass)

log_with_nowtime_str('Original archive uploaded!')
# create additional folder for provider if friday
if today_weekday == 4:
    selenium_with_nextcloud.create_folder(
        url=nextcloud_provider_url,
        folders_name_list=[day_after_tomorrow, two_day_after_tomorrow]
        )
else: pass


# create 'tommorow' folder for nextcloud_upload_for_employee_url
if today_weekday == 4:
    folder_to_create = two_day_after_tomorrow
else:
    folder_to_create = tomorrow
selenium_with_nextcloud.create_folder_and_sub_folder(url=secret_keys.nextcloud_upload_for_employee_url,
                                                     folder_name=folder_to_create,
                                                     subfolder_name='Дата записи',
                                                     password=secret_keys.nextcloud_upload_for_employee_url_pass)
############################################################ my SORT block ########################################################

log_with_nowtime_str('START work with local files in "temp" dir')

unzip_and_sort.unpack_zipfile(dir_with_archive=config.temp_dir,
                              extract_dir=config.temp_dir,
                              encoding='cp866')
print('unpack_zipfile done!!!')

unpacked_dict, unpacked_list = unzip_and_sort.unpacked_tree(config.temp_dir)
unzip_and_sort.find_nested_zip(unpacked_dict,
                               config.temp_dir)
print('find_nested_zip done!!!')

unpacked_dict, unpacked_list = unzip_and_sort.unpacked_tree(config.temp_dir)
unzip_and_sort.rename_files_by_folder(unpacked_list)
print('rename_files_by_folder done!!!')

unpacked_dict, unpacked_list = unzip_and_sort.unpacked_tree(config.temp_dir)
unzip_and_sort.sort_files(original_folder_name=original_folder_name,
                          extract_dir=config.temp_dir,
                           unpacked_list=unpacked_list,
                           not_table_file_types=not_table_file_types,
                           table_file_types=table_file_types,
                           tables_dir=config.tables_dir,
                           not_table_kjz_dir=config.not_table_kjz_dir,
                           others_files_dir=config.others_files_dir)
print('all unzip_and_sort done!!!')
########################################### end my sorting ######################################################################

############################################# xls and ods to xlsx  #####################################################
def close_excel_by_force(excel):
    import win32process
    import win32gui
    import win32api
    import win32con

    # Get the window's process id's
    hwnd = excel.Hwnd
    t, p = win32process.GetWindowThreadProcessId(hwnd)
    # Ask window nicely to close
    win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
    # Allow some time for app to close
    sleep(10)
    # If the application didn't close, force close
    try:
        handle = win32api.OpenProcess(win32con.PROCESS_TERMINATE, 0, p)
        if handle:
            win32api.TerminateProcess(handle, 0)
            win32api.CloseHandle(handle)
    except:
        pass

# rewrite below block to func - cant understand how
mass = []
ffolder = os.listdir(config.tables_dir)
for f in ffolder:
    file = config.tables_dir + '\\' + f
    if (".xlsx" not in f) and ((".xls" in f) or (".ods" in f)):
        mass.append(file)
        res = "." + f.split('.')[-1]
        filename = f.replace(res, "")
        print(filename)

        excel = win32com.client.Dispatch('Excel.Application')
        wb = excel.Workbooks.Open(file)
        excel.DisplayAlerts = False

        wb.SaveAs(config.tables_dir+"\\MOD-"+filename+".xlsx", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension #FileFormat = 56 is for .xls extension
        wb.Close()
        del wb
        excel.Application.Quit()
        excel.Quit()
        close_excel_by_force(excel)
        del excel

sleep(2)
for i in mass:
    os.chmod(i, stat.S_IWRITE)
    os.remove(i)
##########################################################################################################################
log_with_nowtime_str('All tables transform to .xlsx')


################################################ sort files in tables ################################################
# новая проблема с этим - файлы, в которых хэдер - первая строка (обычно это 2-ая)
"""
# appendix block
################################ divide_versions #######################################
for f in os.listdir(config.control_mo_tables_dir):
    file = config.control_mo_tables_dir + '\\' + f
    os.chmod(file, stat.S_IWRITE)
    if ".xlsx" in f.lower():
        divide_versions.divide_replacer(config.control_mo_tables_dir,
                                        f,
                                        divide_versions.divide_check_headers(config.control_mo_tables_dir,
                                                                             f))
########################################################################################
"""
################################################ old sorter - check it  ######################################
for file in os.listdir(config.tables_dir):
    method = unzip_and_sort.check_headers(config.tables_dir, file, dictionary)
    unzip_and_sort.die()
    unzip_and_sort.replacer(config.tables_dir, config.temp_dir, file, method)
shutil.rmtree(config.tables_dir) # delete 'tables_dir'
print('repalacer DONE!', datetime.now().strftime("%H.%M-%d.%m.%Y"))

shutil.make_archive(f'{os.path.join(config.sort_dir, nowtime_str())} - sorted', 'zip', config.temp_dir)
#############################################################################################################
log_with_nowtime_str('END sort local files in temp. Sorted archive saved in local!')


###################################### upload dif tables, img, kjz files and files for colleagues #####################################################
# sorted_files = f'{os.path.join(config.sort_dir, nowtime_str())} - sorted.zip'

if len(os.listdir(config.not_table_kjz_dir)) != 0:
    upload_not_table_kjz = [os.path.join(config.not_table_kjz_dir, file) for file in
                                os.listdir(config.not_table_kjz_dir)]
    try:
        selenium_with_nextcloud.create_folder(f'{secret_keys.nextcloud_upload_for_employee_url}%2F{today}',
                                              ['КЖЗ'],
                                              secret_keys.nextcloud_upload_for_employee_url_pass)
        selenium_with_nextcloud.upload(
            f'{secret_keys.nextcloud_upload_for_employee_url}%2F{today}%2F%D0%9A%D0%96%D0%97',
            upload_not_table_kjz,
            secret_keys.nextcloud_upload_for_employee_url_pass)
    except: print('Somthing wrong in uploading KJZ pics!!!')

# pics to Control MO
archive_other_files_to_cloud = config.others_files_dir
upload_other_types_files = [os.path.join(archive_other_files_to_cloud, file) for file in os.listdir(archive_other_files_to_cloud)]
selenium_with_nextcloud.upload(f'{secret_keys.nextcloud_upload_for_employee_url}%2F{today}',
                               upload_other_types_files,
                               secret_keys.nextcloud_upload_for_employee_url_pass)

# tables to New patients
if len(os.listdir(config.new_patients_tables_dir)) != 0:
    archive_new_patients_tables_to_cloud = config.new_patients_tables_dir
    upload_new_patients_files = [os.path.join(archive_new_patients_tables_to_cloud, file) for file in os.listdir(archive_new_patients_tables_to_cloud)]
    selenium_with_nextcloud.upload(secret_keys.nextcloud_new_patients_form_provider,
                                   upload_new_patients_files)
else: pass

# tables to Delta
if len(os.listdir(config.delta_tables_dir)) != 0:
    archive_delta_tables_to_cloud = config.delta_tables_dir
    upload_delta_files = [os.path.join(archive_delta_tables_to_cloud, file) for file in os.listdir(archive_delta_tables_to_cloud)]
    selenium_with_nextcloud.upload(secret_keys.nextcloud_delta_form_provider,
                                   upload_delta_files)
else: pass
#######################################################################################################################################
log_with_nowtime_str('UPLOAD files to clouds complited!')


################################ gluer(uped-gluer5) #######################################
log_with_nowtime_str('START Base gluer!')
#  TODO: rewrite this crime
def determinate_headers(row):
    header = {}
    for n, i in enumerate(row):
        header[i] = n
    return header

# vars
ffolder = os.listdir(config.control_mo_tables_dir) # be careful - variable 'ffolder' was reassignment!!!
now = datetime.now().date()
mass = []
mass2 = []
errors = []
empty_responce = []

log_with_nowtime_str('Read and glue answers.')
for f in ffolder:
    if ("Свод" not in f) and ((".xlsx" in f) or (".XLSX" in f) or (".xlsm" in f)):
        print(f)
        try:
            wb = openpyxl.load_workbook(config.control_mo_tables_dir + '\\' + f)
            ws = wb.active
            for n, i in enumerate(ws):
                row = []
                for j in i:
                    row.append(j.value)
                if n == 1:
                    hdr = determinate_headers(row)
                elif n > 1:
                    #print(row[hdr["Дата направления ответа от МО в Дирекцию МГКР"]], row[hdr["Ответ от МО"]])
                    rowmod = [[row[hdr["Дата направления запроса в МО"]]
                                  , row[hdr["МО"]]
                                  , row[hdr["Полис ОМС"]]
                                  , row[hdr["Статус"]]
                                  , row[hdr["Подстатус"]]
                                  , row[hdr["Комментарий персонального помощника"]]]]
                    if row[hdr["Ответ от МО"]] not in (None, "", 0):
                        if row[hdr["Дата направления ответа от МО в Дирекцию МГКР"]] in (None, "", 0):
                            rowmod += [[now, row[hdr["Ответ от МО"]]]]
                        else:
                            rowmod += [[row[hdr["Дата направления ответа от МО в Дирекцию МГКР"]], row[hdr["Ответ от МО"]]]]
                        rowmod += [False]
                        #print(rowmod)
                        if rowmod not in mass2:
                            mass2.append(rowmod)
                            mass.append(rowmod + [[str(f)]])
                    else:
                        if row[hdr["Дата направления ответа от МО в Дирекцию МГКР"]] not in (None, "", 0):
                            rowmod = rowmod[0] + [row[hdr["Дата направления ответа от МО в Дирекцию МГКР"]], row[hdr["Ответ от МО"]]]
                            empty_responce.append(rowmod + [str(f)])
        except Exception as err:
            errors.append([f, str(err)])
            print(f, err)

log_with_nowtime_str('Adding answers to file with questions.')
overmass = []
for _ in os.listdir(config.main_dir):
    if _.startswith('Свод'):
        print(f"{config.main_dir}\\{_}")
        wb = openpyxl.load_workbook(f"{config.main_dir}\\{_}") # переписать, так, чтобы брал из иной папки файл совода
del wb['Ответы вне реестра'], wb['Другие ответы'], wb['Пустые ответы'], wb['Ошибки чтения файлов']  # check it del function

ws = wb.active
ws.protection.disable()
for n, i in enumerate(ws):
    row = []
    for j in i:
        row.append(j.value)
    if n == 1:
        hdr = determinate_headers(row)
    elif n > 1:
        for n2, m in enumerate(mass):
            check = [row[hdr["Дата направления запроса в МО"]]
                , row[hdr["МО"]]
                , row[hdr["Полис ОМС"]]
                , row[hdr["Статус"]]
                , row[hdr["Подстатус"]]
                , row[hdr["Комментарий персонального помощника"]]]
            if check == m[0]: #если есть ответ от МО
                mass[n2][2] = True
                if (row[hdr["Дата направления ответа от МО в Дирекцию МГКР"]] not in (None, "", 0)) or (row[hdr["Ответ от МО"]] not in (None, "", 0)):
                    # если в своде уже есть ответ
                    if [row[hdr["Дата направления ответа от МО в Дирекцию МГКР"]], row[hdr["Ответ от МО"]]] != m[1]:
                        overmass.append(m[0]+m[1]+m[3]) #другой ответ
                    else:
                        pass #дубль
                else: # если в своде ответа нет, то обновляем свод
                    ws.cell(n+1, hdr["Дата направления ответа от МО в Дирекцию МГКР"]+1).value = m[1][0]
                    ws.cell(n+1, hdr["Ответ от МО"]+1).value = m[1][1]
                    row[hdr["Дата направления ответа от МО в Дирекцию МГКР"]] = m[1][0]
                    row[hdr["Ответ от МО"]] = m[1][1]


header = ["Дата направления запроса в МО",
          "МО", "Полис ОМС",
          "Статус", "Подстатус",
          "Комментарий персонального помощника",
          "Дата направления ответа от МО в Дирекцию МГКР",
          "Ответ от МО", "Файл"]

ws2 = wb.create_sheet("Ответы вне реестра")
ws2.append(header)
for m in mass:
    if m[2] == False:
        ws2.append(m[0]+m[1]+m[3])
ws3 = wb.create_sheet("Другие ответы")
ws3.append(header)
for o in overmass:
    ws3.append(o)
ws4 = wb.create_sheet("Пустые ответы")
ws4.append(header)
for emp in empty_responce:
    ws4.append(emp)
ws5 = wb.create_sheet("Ошибки чтения файлов")
ws5.append(["Файл", "Ошибка"])
for e in errors:
    ws5.append(e)

if len(overmass) > 0:
    print('Проверьте - есть "Другие ответы"')
if len(overmass) > 0:
    print('Проверьте - есть "Пустые ответы"')
if len(overmass) > 0:
    print('Проверьте - есть "Ошибки чтения файлов"')

for _ in [config.control_mo_tables_dir,
          config.delta_tables_dir,
          config.new_patients_tables_dir,
          config.not_table_kjz_dir,
          config.others_files_dir]:
    shutil.rmtree(_)



log_with_nowtime_str('Start local SAVE Result.')
wb.save(f"{config.main_dir}\\+Свод.xlsx")
########################################################################################
log_with_nowtime_str('- = ALL COMPLETED!!! THE RESULT WAS SAVED! = -')
telegram_actions.send_message_for_users_by_list('Файлы "Контроль МО" обработаны - произведена загрузка в облако!')
